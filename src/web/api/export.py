"""Device inventory export to CSV and XLSX."""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from src.storage.repository import DeviceRepository
from src.web.config import EXPORT_COLUMNS
from src.web.dependencies import get_repo

router = APIRouter()


def _flatten_device(dev: dict[str, Any]) -> dict[str, str]:
    """Flatten a device dict for export — lists become semicolon-separated strings."""
    row: dict[str, str] = {}
    for col in EXPORT_COLUMNS:
        val = dev.get(col, "")
        if isinstance(val, list):
            row[col] = "; ".join(str(v) for v in val)
        elif val is None:
            row[col] = ""
        else:
            row[col] = str(val)
    return row


@router.get("/api/export/csv")
async def export_csv(
    status: str | None = None,
    source: str | None = None,
    repo: DeviceRepository = Depends(get_repo),
) -> StreamingResponse:
    """Stream the (filtered) device inventory as CSV."""
    devices = repo.get_all_devices(status=status, source=source)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for dev in devices:
        writer.writerow(_flatten_device(dev))
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"device_inventory_{ts}.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/export/xlsx")
async def export_xlsx(
    status: str | None = None,
    source: str | None = None,
    repo: DeviceRepository = Depends(get_repo),
) -> StreamingResponse:
    """Stream the (filtered) device inventory as a styled XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    devices = repo.get_all_devices(status=status, source=source)

    wb = Workbook()
    ws = wb.active
    ws.title = "Device Inventory"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    status_colors = {
        "FULLY_MANAGED": "C6EFCE",
        "MANAGED": "DFF0D8",
        "NO_EDR": "FFC7CE",
        "NO_MDM": "FFEB9C",
        "IDP_ONLY": "FFD699",
        "STALE": "D9D9D9",
        "UNKNOWN": "F2F2F2",
    }
    for row_idx, dev in enumerate(devices, 2):
        flat = _flatten_device(dev)
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=flat[col_name])
            if col_name == "status" and flat[col_name] in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[flat[col_name]],
                    end_color=status_colors[flat[col_name]],
                    fill_type="solid",
                )

    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(devices) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"device_inventory_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
