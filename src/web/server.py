from __future__ import annotations

import csv
import hashlib
import io
import os
import secrets
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, AsyncGenerator

import jwt
from fastapi import BackgroundTasks, Cookie, FastAPI, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from src.insights import generate_insights, generate_report_text
from src.storage.repository import DeviceRepository
from src.sync_engine import SyncEngine

DB_PATH = os.getenv("DB_PATH", "data/devices.db")
APP_URL = os.getenv("APP_URL", "http://localhost:8080")
AUTH_USERNAME = os.getenv("AUTH_USERNAME", "admin")
AUTH_PASSWORD = os.getenv("AUTH_PASSWORD", "")
JWT_SECRET = os.getenv("JWT_SECRET", secrets.token_hex(32))
JWT_EXPIRY_HOURS = 24
DIST_DIR = Path(__file__).resolve().parents[2] / "frontend" / "dist"
_IS_HTTPS = APP_URL.startswith("https://")

# Okta OIDC
OKTA_OIDC_ISSUER = os.getenv("OKTA_OIDC_ISSUER", "")
OKTA_OIDC_CLIENT_ID = os.getenv("OKTA_OIDC_CLIENT_ID", "")
OKTA_OIDC_CLIENT_SECRET = os.getenv("OKTA_OIDC_CLIENT_SECRET", "")
_OKTA_OIDC_ENABLED = bool(OKTA_OIDC_ISSUER and OKTA_OIDC_CLIENT_ID and OKTA_OIDC_CLIENT_SECRET)


@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncGenerator[None, None]:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.interval import IntervalTrigger

    sync_interval = int(os.getenv("SYNC_INTERVAL_HOURS", "6"))
    sync_on_startup = os.getenv("SYNC_ON_STARTUP", "true").lower() == "true"

    scheduler = BackgroundScheduler()
    engine = SyncEngine(DB_PATH)

    def _job() -> None:
        global _syncing
        _syncing = True
        try:
            engine.run()
            refresh_cache()
        except Exception:
            pass
        finally:
            _syncing = False

    scheduler.add_job(
        _job,
        trigger=IntervalTrigger(hours=sync_interval),
        id="device_sync",
        replace_existing=True,
    )
    scheduler.start()
    if sync_on_startup:
        if SyncEngine.should_skip_startup_sync(DB_PATH):
            from structlog import get_logger
            get_logger(__name__).info("startup_sync_skipped", reason="last_sync_within_2h")
            refresh_cache()
        else:
            # Run startup sync in background — don't block server startup
            import threading
            from structlog import get_logger
            get_logger(__name__).info("startup_sync_background")
            threading.Thread(target=_job, daemon=True).start()
    else:
        refresh_cache()
    yield
    scheduler.shutdown()


app = FastAPI(title="Klar Device Normalizer", lifespan=lifespan)

# ── In-memory cache — refreshed after each sync ──────────────────────────
_cache: dict[str, Any] = {}
_syncing = False


def refresh_cache() -> None:
    """Pre-compute expensive data so API responses are instant."""
    try:
        repo = DeviceRepository(DB_PATH)
        devices = repo.get_all_devices()
        summary = repo.get_summary()
        by_status = summary.get("by_status", {})
        total_devices = summary.get("total", 0)
        if total_devices > 0:
            weighted = sum(by_status.get(s, 0) * w for s, w in RISK_WEIGHTS.items())
            summary["risk_score"] = round(weighted / total_devices, 1)
        else:
            summary["risk_score"] = 0
        history = repo.get_status_history(limit=30)
        prev = repo.get_previous_snapshot()
        trends: dict[str, int] = {}
        if prev:
            for status_key, col_name in [("FULLY_MANAGED", "fully_managed"), ("MANAGED", "managed"),
                ("NO_EDR", "no_edr"), ("NO_MDM", "no_mdm"), ("IDP_ONLY", "idp_only"),
                ("STALE", "stale"), ("SERVER", "server")]:
                trends[status_key] = by_status.get(status_key, 0) - prev.get(col_name, 0)

        _cache["summary"] = summary
        _cache["trends"] = {"trends": trends, "has_previous": prev is not None}
        _cache["history"] = {"history": history}
        _cache["insights"] = {"actions": generate_insights(devices, summary, history)}
        from structlog import get_logger
        get_logger(__name__).info("cache_refreshed")
    except Exception as exc:
        from structlog import get_logger
        get_logger(__name__).warning("cache_refresh_failed", error=str(exc))


# Serve static assets (JS/CSS bundles)
if (DIST_DIR / "assets").exists():
    app.mount("/assets", StaticFiles(directory=DIST_DIR / "assets"), name="assets")


def _get_repo() -> DeviceRepository:
    return DeviceRepository(DB_PATH)


# ── Auth ─────────────────────────────────────────────────────────────────────

PUBLIC_PATHS = {"/auth/login", "/auth/logout", "/auth/okta", "/auth/okta/callback", "/auth/me", "/favicon.svg", "/healthz", "/api/version"}


def _create_token(username: str) -> str:
    return jwt.encode(
        {"sub": username, "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS)},
        JWT_SECRET, algorithm="HS256",
    )


def _verify_token(token: str | None) -> str | None:
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        return payload.get("sub")
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        return None


@app.middleware("http")
async def auth_middleware(request: Request, call_next: Any) -> Any:
    path = request.url.path

    # Skip auth only if no auth method is configured at all
    if not AUTH_PASSWORD and not _OKTA_OIDC_ENABLED:
        return await call_next(request)
    if path in PUBLIC_PATHS or path.startswith("/assets/"):
        return await call_next(request)

    token = request.cookies.get("klar_session")
    user = _verify_token(token)

    if not user:
        # API calls get 401, page requests get redirect to login
        if path.startswith("/api/"):
            return JSONResponse({"detail": "Unauthorized"}, status_code=401)
        return HTMLResponse(content=_login_page(), status_code=200)

    return await call_next(request)


class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/auth/login")
async def auth_login(body: LoginRequest) -> Any:
    if not AUTH_PASSWORD:
        return JSONResponse({"error": "Auth not configured"}, status_code=400)

    # Constant-time comparison
    user_ok = secrets.compare_digest(body.username, AUTH_USERNAME)
    pass_ok = secrets.compare_digest(body.password, AUTH_PASSWORD)

    if not (user_ok and pass_ok):
        return JSONResponse({"error": "Invalid credentials"}, status_code=401)

    token = _create_token(body.username)
    response = JSONResponse({"ok": True, "user": body.username})
    response.set_cookie(
        key="klar_session",
        value=token,
        httponly=True,
        secure=_IS_HTTPS,
        samesite="lax",
        max_age=JWT_EXPIRY_HOURS * 3600,
        path="/",
    )
    return response


@app.get("/auth/me")
async def auth_me(request: Request) -> Any:
    token = request.cookies.get("klar_session")
    user = _verify_token(token)
    return JSONResponse({"user": user or "unknown"})


@app.get("/auth/logout")
async def auth_logout() -> Any:
    from fastapi.responses import RedirectResponse
    response = RedirectResponse("/")
    response.delete_cookie("klar_session", path="/")
    return response


# ── Okta OIDC ────────────────────────────────────────────────────────────

@app.get("/auth/okta")
async def auth_okta_redirect() -> Any:
    """Redirect user to Okta for authentication."""
    if not _OKTA_OIDC_ENABLED:
        return JSONResponse({"error": "Okta OIDC not configured"}, status_code=400)
    import urllib.parse

    # FIX-1: Generate state and store in cookie for CSRF protection
    state = secrets.token_hex(16)
    params = urllib.parse.urlencode({
        "client_id": OKTA_OIDC_CLIENT_ID,
        "response_type": "code",
        "scope": "openid email profile",
        "redirect_uri": f"{APP_URL}/auth/okta/callback",
        "state": state,
    })
    from fastapi.responses import RedirectResponse
    response = RedirectResponse(f"{OKTA_OIDC_ISSUER}/v1/authorize?{params}")
    response.set_cookie(
        key="okta_state", value=state, httponly=True, samesite="lax",
        secure=_IS_HTTPS, max_age=300, path="/",
    )
    return response


@app.get("/auth/okta/callback")
async def auth_okta_callback(code: str = "", error: str = "", state: str = "", request: Request = None) -> Any:
    """Handle Okta OIDC callback — exchange code for token."""
    import html as html_mod

    # FIX-3: Escape all user-controlled values before rendering
    if error:
        safe_error = html_mod.escape(error)
        return HTMLResponse(f"<h3>Login failed: {safe_error}</h3><a href='/'>Try again</a>")
    if not code or not _OKTA_OIDC_ENABLED:
        return HTMLResponse("<h3>Invalid callback</h3><a href='/'>Try again</a>")

    # FIX-1: Validate state parameter against cookie
    expected_state = request.cookies.get("okta_state") if request else None
    if not expected_state or not secrets.compare_digest(state, expected_state):
        return HTMLResponse("<h3>Invalid state — possible CSRF attack</h3><a href='/'>Try again</a>")

    import httpx
    try:
        async with httpx.AsyncClient() as client:
            token_resp = await client.post(
                f"{OKTA_OIDC_ISSUER}/v1/token",
                data={
                    "grant_type": "authorization_code",
                    "code": code,
                    "redirect_uri": f"{APP_URL}/auth/okta/callback",
                    "client_id": OKTA_OIDC_CLIENT_ID,
                    "client_secret": OKTA_OIDC_CLIENT_SECRET,
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            if token_resp.status_code != 200:
                return HTMLResponse("<h3>Token exchange failed</h3><a href='/'>Try again</a>")
            tokens = token_resp.json()

            # FIX-2: Validate ID token claims
            id_token_raw = tokens.get("id_token")
            if id_token_raw:
                try:
                    # Decode without verification first to check claims
                    # (signature is implicitly trusted since we got this token
                    # directly from Okta over TLS in a server-to-server call)
                    claims = jwt.decode(id_token_raw, options={"verify_signature": False})
                    if claims.get("iss") != OKTA_OIDC_ISSUER:
                        return HTMLResponse("<h3>Invalid token issuer</h3><a href='/'>Try again</a>")
                    if claims.get("aud") != OKTA_OIDC_CLIENT_ID:
                        return HTMLResponse("<h3>Invalid token audience</h3><a href='/'>Try again</a>")
                except jwt.InvalidTokenError:
                    return HTMLResponse("<h3>Invalid ID token</h3><a href='/'>Try again</a>")

            # Get user info
            userinfo_resp = await client.get(
                f"{OKTA_OIDC_ISSUER}/v1/userinfo",
                headers={"Authorization": f"Bearer {tokens['access_token']}"},
            )
            if userinfo_resp.status_code != 200:
                return HTMLResponse("<h3>Failed to get user info</h3><a href='/'>Try again</a>")
            userinfo = userinfo_resp.json()
    except Exception:
        # FIX-3: Don't leak exception details
        return HTMLResponse("<h3>Authentication failed</h3><a href='/'>Try again</a>")

    email = userinfo.get("email") or userinfo.get("preferred_username") or ""

    # FIX-5: Validate user email domain
    allowed_domains = os.getenv("OKTA_ALLOWED_DOMAINS", "").strip()
    if allowed_domains:
        domains = [d.strip().lower() for d in allowed_domains.split(",") if d.strip()]
        email_domain = email.split("@")[-1].lower() if "@" in email else ""
        if domains and email_domain not in domains:
            return HTMLResponse("<h3>Access denied — your domain is not authorized</h3><a href='/'>Back</a>")

    if not email:
        return HTMLResponse("<h3>No email in Okta profile</h3><a href='/'>Try again</a>")

    token = _create_token(email)

    from fastapi.responses import RedirectResponse
    response = RedirectResponse("/")
    response.set_cookie(
        key="klar_session", value=token, httponly=True,
        secure=_IS_HTTPS, samesite="lax", max_age=JWT_EXPIRY_HOURS * 3600, path="/",
    )
    response.delete_cookie("okta_state", path="/")
    return response


def _login_page() -> str:
    okta_section = ""
    if _OKTA_OIDC_ENABLED:
        okta_section = """
      <div style="margin-top:20px;text-align:center">
        <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px">
          <div style="flex:1;height:1px;background:rgba(255,255,255,0.08)"></div>
          <span style="font-size:11px;color:var(--gray-4);text-transform:uppercase;letter-spacing:1px">or</span>
          <div style="flex:1;height:1px;background:rgba(255,255,255,0.08)"></div>
        </div>
        <a href="/auth/okta" style="display:flex;align-items:center;justify-content:center;gap:8px;width:100%;padding:12px;background:transparent;border:1.5px solid rgba(255,255,255,0.15);border-radius:10px;color:var(--white);font-size:13px;font-weight:500;text-decoration:none;transition:all 0.2s;font-family:inherit"
           onmouseover="this.style.borderColor='rgba(255,255,255,0.3)';this.style.background='rgba(255,255,255,0.03)'"
           onmouseout="this.style.borderColor='rgba(255,255,255,0.15)';this.style.background='transparent'">
          <svg width="40" height="14" viewBox="0 0 200 70" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M34.8 0C15.6 0 0 15.6 0 34.8s15.6 34.8 34.8 34.8 34.8-15.6 34.8-34.8S54 0 34.8 0zm0 52.2c-9.6 0-17.4-7.8-17.4-17.4s7.8-17.4 17.4-17.4 17.4 7.8 17.4 17.4-7.8 17.4-17.4 17.4zm70.8-36.6h-.6c-4.2 0-7.8 1.2-10.8 3.6V2.4h-15v64.8h15V39c0-5.4 3.6-9 8.4-9 1.2 0 2.4.6 3 .6h.6l4.8-15h-5.4zm29.4 0c-3 0-6 .6-9 1.8l-1.2.6 4.2 12 1.2-.6c1.8-.6 3.6-1.2 5.4-1.2 3 0 4.2 1.2 4.2 3v1.2l-3.6.6c-10.2 1.8-18 6-18 15 0 7.8 5.4 13.2 13.8 13.2 4.8 0 7.8-1.8 9.6-3.6v3h13.8V34.2c0-12.6-7.2-18.6-20.4-18.6zm5.4 36c-1.8 1.2-3.6 1.8-6 1.8-2.4 0-4.2-1.2-4.2-3.6 0-3.6 4.2-5.4 10.2-6.6v8.4z" fill="#007DC1"/></svg>
          Sign in with Okta
        </a>
      </div>"""
    return """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Klar — Sign in</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,400;0,500;0,700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{--black:#0f0f0f;--white:#fff;--gray-1:#f7f7f5;--gray-2:#ebebea;--gray-3:#c4c4c3;--gray-4:#8a8a89;--gray-5:#555;--red:#c0392b;--green:#10b981}
*{margin:0;padding:0;box-sizing:border-box}
html,body{height:100%}
body{font-family:'DM Sans',system-ui,sans-serif;background:var(--black);color:var(--white);overflow:hidden}

/* Animated grid background */
.grid-bg{position:fixed;inset:0;z-index:0;
  background-image:
    linear-gradient(rgba(255,255,255,0.03) 1px,transparent 1px),
    linear-gradient(90deg,rgba(255,255,255,0.03) 1px,transparent 1px);
  background-size:48px 48px;
  animation:gridShift 20s linear infinite}
@keyframes gridShift{from{transform:translate(0,0)}to{transform:translate(48px,48px)}}

/* Scan line effect */
.scanline{position:fixed;inset:0;z-index:1;pointer-events:none;
  background:repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(255,255,255,0.008) 2px,rgba(255,255,255,0.008) 4px)}

/* Floating orb */
.orb{position:fixed;width:600px;height:600px;border-radius:50%;z-index:0;filter:blur(120px);opacity:0.07;
  background:radial-gradient(circle,var(--green),transparent 70%);
  top:-200px;right:-200px;animation:orbFloat 15s ease-in-out infinite alternate}
@keyframes orbFloat{0%{transform:translate(0,0)}100%{transform:translate(-80px,80px)}}

/* Layout */
.wrapper{position:relative;z-index:2;display:flex;height:100vh}
.left{flex:1;display:flex;flex-direction:column;justify-content:center;padding:0 8vw}
.right{width:480px;display:flex;align-items:center;justify-content:center;padding:40px}

/* Left panel */
.brand{font-size:clamp(42px,5vw,64px);font-weight:700;letter-spacing:-2px;line-height:1;margin-bottom:12px;
  animation:fadeUp 0.8s ease both}
.tagline{font-size:15px;color:var(--gray-4);font-weight:400;line-height:1.6;max-width:340px;
  animation:fadeUp 0.8s ease 0.1s both}
.tagline strong{color:var(--gray-3);font-weight:500}
.status-bar{position:absolute;bottom:40px;left:8vw;display:flex;gap:24px;
  animation:fadeUp 0.8s ease 0.3s both}
.status-item{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--gray-4);
  display:flex;align-items:center;gap:6px}
.dot{width:6px;height:6px;border-radius:50%;background:var(--green);
  box-shadow:0 0 8px var(--green);animation:pulse 2s ease-in-out infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.4}}

/* Right panel — card */
.card{width:100%;max-width:360px;background:rgba(255,255,255,0.03);
  border:1px solid rgba(255,255,255,0.08);border-radius:20px;padding:44px 36px;
  backdrop-filter:blur(20px);animation:fadeUp 0.6s ease 0.15s both}
.card-title{font-size:14px;font-weight:500;color:var(--gray-3);margin-bottom:32px;
  letter-spacing:0.5px;text-transform:uppercase}
.field{margin-bottom:22px}
.field label{display:block;font-size:11px;font-weight:500;color:var(--gray-4);
  text-transform:uppercase;letter-spacing:1.2px;margin-bottom:8px}
.field input{width:100%;padding:12px 16px;background:rgba(255,255,255,0.04);
  border:1px solid rgba(255,255,255,0.1);border-radius:10px;color:var(--white);
  font-size:14px;font-family:'DM Sans',sans-serif;outline:none;transition:all 0.25s}
.field input:focus{border-color:rgba(255,255,255,0.3);background:rgba(255,255,255,0.06);
  box-shadow:0 0 0 3px rgba(255,255,255,0.04)}
.field input::placeholder{color:var(--gray-4)}
.btn{width:100%;padding:14px;background:var(--white);color:var(--black);border:none;
  border-radius:10px;font-size:14px;font-weight:600;font-family:inherit;cursor:pointer;
  transition:all 0.2s;letter-spacing:0.2px;margin-top:4px}
.btn:hover{background:var(--gray-2);transform:translateY(-1px);box-shadow:0 4px 20px rgba(255,255,255,0.1)}
.btn:active{transform:translateY(0)}
.btn:disabled{opacity:0.5;cursor:not-allowed;transform:none}
.error-msg{color:var(--red);font-size:12px;margin-bottom:16px;padding:10px 14px;
  background:rgba(192,57,43,0.08);border:1px solid rgba(192,57,43,0.15);border-radius:8px;
  display:none;animation:shake 0.4s ease}
@keyframes shake{0%,100%{transform:translateX(0)}20%,60%{transform:translateX(-6px)}40%,80%{transform:translateX(6px)}}
.footer{text-align:center;margin-top:24px;font-size:11px;color:var(--gray-4)}

@keyframes fadeUp{from{opacity:0;transform:translateY(16px)}to{opacity:1;transform:translateY(0)}}

/* Responsive */
@media(max-width:900px){
  .wrapper{flex-direction:column}
  .left{flex:none;padding:40px 32px 20px;text-align:center;align-items:center}
  .left .tagline{margin:0 auto}
  .right{width:100%;flex:1;padding:20px 32px 40px}
  .status-bar{position:static;justify-content:center;margin-top:20px}
}
</style>
</head>
<body>
<div class="grid-bg"></div>
<div class="scanline"></div>
<div class="orb"></div>

<div class="wrapper">
  <div class="left">
    <div class="brand">Klar</div>
    <div class="tagline">
      <strong>Device Normalizer</strong><br>
      Fleet visibility across JumpCloud, CrowdStrike &amp; Okta — unified in one secure dashboard.
    </div>
    <div class="status-bar">
      <div class="status-item"><span class="dot"></span> System operational</div>
      <div class="status-item" id="ver">loading...</div>
    </div>
  </div>

  <div class="right">
    <div class="card">
      <div class="card-title">Sign in to continue</div>
      <div class="error-msg" id="err"></div>
      <form id="form" autocomplete="on">
        <div class="field">
          <label for="user">Username</label>
          <input type="text" id="user" name="username" autocomplete="username" placeholder="admin" required>
        </div>
        <div class="field">
          <label for="pass">Password</label>
          <input type="password" id="pass" name="password" autocomplete="current-password" placeholder="&bull;&bull;&bull;&bull;&bull;&bull;&bull;&bull;" required>
        </div>
        <button class="btn" type="submit" id="btn">Sign in</button>
      </form>
      {okta_section}
      <div class="footer">Klar — IT Security Team</div>
    </div>
  </div>
</div>

<script>
document.getElementById('form').addEventListener('submit',async(e)=>{
  e.preventDefault();
  const err=document.getElementById('err'),btn=document.getElementById('btn');
  err.style.display='none';
  btn.disabled=true;btn.textContent='Authenticating...';
  try{
    const r=await fetch('/auth/login',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({username:document.getElementById('user').value,
        password:document.getElementById('pass').value})});
    if(r.ok){btn.textContent='Redirecting...';location.reload()}
    else{const d=await r.json();err.textContent=d.error||'Authentication failed';
      err.style.display='block';btn.disabled=false;btn.textContent='Sign in';
      document.getElementById('pass').value='';document.getElementById('pass').focus()}
  }catch(ex){err.textContent='Connection error';err.style.display='block';
    btn.disabled=false;btn.textContent='Sign in'}
});
fetch('/api/version').then(r=>r.json()).then(d=>{
  const v=d.version==='dev'?'dev':d.version.slice(0,7);
  document.getElementById('ver').textContent=v;
}).catch(()=>{document.getElementById('ver').textContent='v1.0'});
</script>
</body>
</html>""".replace("{okta_section}", okta_section)


# ── Health check (public, no auth) ────────────────────────────────────────────

APP_VERSION = os.getenv("APP_VERSION", "dev")
APP_BUILD_DATE = os.getenv("APP_BUILD_DATE", "")


@app.get("/healthz")
async def healthz() -> Any:
    return JSONResponse(content={"status": "ok", "syncing": _syncing, "version": APP_VERSION})


@app.get("/api/version")
async def api_version() -> Any:
    return JSONResponse(content={"version": APP_VERSION, "build_date": APP_BUILD_DATE})


# ── API Routes ───────────────────────────────────────────────────────────────

@app.get("/api/devices")
async def api_devices(status: str | None = None, source: str | None = None) -> Any:
    repo = _get_repo()
    devices = repo.get_all_devices(status=status, source=source)
    # Mark acknowledged devices
    acked = repo.get_acknowledged_details()
    for d in devices:
        cid = d.get("canonical_id", "")
        if cid in acked:
            d["acknowledged"] = True
            d["ack_reason"] = acked[cid]["reason"]
            d["ack_by"] = acked[cid]["by"]
            d["ack_at"] = acked[cid]["at"]
        else:
            d["acknowledged"] = False
    return JSONResponse(content={"devices": devices})


class AckRequest(BaseModel):
    reason: str = ""
    by: str = ""


@app.post("/api/devices/{canonical_id}/ack")
async def ack_device(canonical_id: str, body: AckRequest, request: Request) -> Any:
    # Use logged-in user if 'by' not provided
    ack_by = body.by
    if not ack_by:
        token = request.cookies.get("klar_session")
        ack_by = _verify_token(token) or "unknown"
    repo = _get_repo()
    repo.acknowledge_device(canonical_id, reason=body.reason, by=ack_by)
    return JSONResponse(content={"ok": True, "canonical_id": canonical_id, "by": ack_by})


@app.delete("/api/devices/{canonical_id}/ack")
async def unack_device(canonical_id: str) -> Any:
    repo = _get_repo()
    repo.unacknowledge_device(canonical_id)
    return JSONResponse(content={"ok": True, "canonical_id": canonical_id})


RISK_WEIGHTS: dict[str, int] = {
    "FULLY_MANAGED": 100,
    "MANAGED": 80,
    "SERVER": 75,
    "NO_MDM": 40,
    "NO_EDR": 25,
    "IDP_ONLY": 15,
    "STALE": 5,
    "UNKNOWN": 10,
}


@app.get("/api/summary")
async def api_summary() -> Any:
    repo = _get_repo()
    summary = repo.get_summary()
    # Calculate risk score
    by_status = summary.get("by_status", {})
    total = summary.get("total", 0)
    if total > 0:
        weighted = sum(by_status.get(s, 0) * w for s, w in RISK_WEIGHTS.items())
        score = round(weighted / total, 1)
    else:
        score = 0
    summary["risk_score"] = score
    summary["syncing"] = _syncing
    # Next sync estimate
    last_sync = repo.get_last_sync_run()
    if last_sync and last_sync.get("finished_at"):
        try:
            finished = datetime.fromisoformat(str(last_sync["finished_at"]).replace("Z", "+00:00"))
            if finished.tzinfo is None:
                finished = finished.replace(tzinfo=timezone.utc)
            interval_h = int(os.getenv("SYNC_INTERVAL_HOURS", "6"))
            next_sync = finished + timedelta(hours=interval_h)
            summary["next_sync"] = next_sync.isoformat()
            summary["sync_interval_hours"] = interval_h
        except Exception:
            pass
    return JSONResponse(content=summary)


@app.get("/api/diff")
async def api_diff() -> Any:
    """Changes between the last two syncs."""
    repo = _get_repo()
    devices = repo.get_all_devices()
    new_devices = repo.get_new_devices()
    disappeared = repo.get_recently_deleted()
    newly_stale = repo.get_newly_stale()

    # Status changes: compare current snapshot vs previous
    history = repo.get_status_history(limit=2)
    status_changes: dict[str, dict[str, int]] = {}
    if len(history) >= 2:
        curr, prev = history[-1], history[-2]
        for col in ["fully_managed", "managed", "no_edr", "no_mdm", "idp_only", "stale", "server"]:
            status_key = col.upper()
            c_val = curr.get(col, 0)
            p_val = prev.get(col, 0)
            if c_val != p_val:
                status_changes[status_key] = {"previous": p_val, "current": c_val, "delta": c_val - p_val}

    def _dev_summary(d: dict) -> dict:
        return {
            "hostname": (d.get("hostnames") or ["?"])[0],
            "owner": d.get("owner_email"),
            "status": d.get("status"),
            "sources": d.get("sources", []),
        }

    return JSONResponse(content={
        "new_devices": {"count": len(new_devices), "devices": [_dev_summary(d) for d in new_devices[:20]]},
        "disappeared": {"count": len(disappeared), "devices": [_dev_summary(d) for d in disappeared[:20]]},
        "newly_stale": {"count": len(newly_stale), "devices": [_dev_summary(d) for d in newly_stale[:10]]},
        "status_changes": status_changes,
        "total_current": len(devices),
    })


@app.get("/api/history")
async def api_history(limit: int = 30) -> Any:
    if "history" in _cache:
        return JSONResponse(content=_cache["history"])
    repo = _get_repo()
    history = repo.get_status_history(limit=limit)
    return JSONResponse(content={"history": history})


@app.get("/api/trends")
async def api_trends() -> Any:
    if "trends" in _cache:
        return JSONResponse(content=_cache["trends"])
    repo = _get_repo()
    prev = repo.get_previous_snapshot()
    summary = repo.get_summary()
    current = summary.get("by_status", {})
    trends: dict[str, int] = {}
    if prev:
        for status_key, col_name in [
            ("FULLY_MANAGED", "fully_managed"),
            ("MANAGED", "managed"),
            ("NO_EDR", "no_edr"),
            ("NO_MDM", "no_mdm"),
            ("IDP_ONLY", "idp_only"),
            ("STALE", "stale"),
            ("SERVER", "server"),
        ]:
            old_val = prev.get(col_name, 0)
            new_val = current.get(status_key, 0)
            trends[status_key] = new_val - old_val
    return JSONResponse(content={"trends": trends, "has_previous": prev is not None})


@app.get("/api/sync/last")
async def api_sync_last() -> Any:
    repo = _get_repo()
    last = repo.get_last_sync_run()
    return JSONResponse(content={"last_sync": last})


EXPORT_COLUMNS = [
    "canonical_id", "hostnames", "serial_number", "owner_email", "owner_name",
    "os_type", "status", "sources", "coverage_gaps", "confidence_score",
    "match_reason", "days_since_seen", "first_seen", "last_seen",
]


def _flatten_device(dev: dict[str, Any]) -> dict[str, str]:
    """Flatten a device dict for export — lists become semicolon-separated strings."""
    row: dict[str, str] = {}
    for col in EXPORT_COLUMNS:
        val = dev.get(col, "")
        if isinstance(val, list):
            row[col] = "; ".join(str(v) for v in val)
        elif val is None:
            row[col] = ""
        else:
            row[col] = str(val)
    return row


@app.get("/api/export/csv")
async def export_csv(status: str | None = None, source: str | None = None) -> StreamingResponse:
    repo = _get_repo()
    devices = repo.get_all_devices(status=status, source=source)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for dev in devices:
        writer.writerow(_flatten_device(dev))
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"device_inventory_{ts}.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/export/xlsx")
async def export_xlsx(status: str | None = None, source: str | None = None) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    repo = _get_repo()
    devices = repo.get_all_devices(status=status, source=source)

    wb = Workbook()
    ws = wb.active
    ws.title = "Device Inventory"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    status_colors = {
        "FULLY_MANAGED": "C6EFCE",
        "MANAGED": "DFF0D8",
        "NO_EDR": "FFC7CE",
        "NO_MDM": "FFEB9C",
        "IDP_ONLY": "FFD699",
        "STALE": "D9D9D9",
        "UNKNOWN": "F2F2F2",
    }
    for row_idx, dev in enumerate(devices, 2):
        flat = _flatten_device(dev)
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=flat[col_name])
            if col_name == "status" and flat[col_name] in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[flat[col_name]],
                    end_color=status_colors[flat[col_name]],
                    fill_type="solid",
                )

    # Auto-width columns
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(devices) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"device_inventory_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Coverage Gaps ───────────────────────────────────────────────────────────

@app.get("/api/gaps")
async def api_gaps() -> Any:
    repo = _get_repo()
    devices = repo.get_all_devices()
    gaps: dict[str, list[dict[str, Any]]] = {
        "missing_edr": [],
        "missing_mdm": [],
        "missing_idp": [],
    }
    for dev in devices:
        dev_gaps = dev.get("coverage_gaps", [])
        summary = {
            "canonical_id": dev.get("canonical_id"),
            "hostnames": dev.get("hostnames", []),
            "owner_email": dev.get("owner_email"),
            "status": dev.get("status"),
            "sources": dev.get("sources", []),
            "days_since_seen": dev.get("days_since_seen"),
        }
        for gap in dev_gaps:
            if gap in gaps:
                gaps[gap].append(summary)
    return JSONResponse(content={
        "gaps": {k: v for k, v in gaps.items()},
        "counts": {k: len(v) for k, v in gaps.items()},
    })


@app.get("/api/people")
async def api_people() -> Any:
    """Person-centric view: email → devices → compliance. Includes unassigned."""
    repo = _get_repo()
    devices = repo.get_all_devices()
    acked = repo.get_acknowledged()

    by_owner: dict[str, list[dict[str, Any]]] = {}
    for d in devices:
        email = d.get("owner_email") or "unassigned"
        by_owner.setdefault(email.lower(), []).append(d)

    people = []
    for email, devs in sorted(by_owner.items()):
        non_acked = [d for d in devs if d.get("canonical_id") not in acked]
        managed_count = sum(1 for d in non_acked if d.get("status") in ("MANAGED", "FULLY_MANAGED"))
        total_count = len(non_acked)
        has_edr = any("crowdstrike" in d.get("sources", []) for d in non_acked)
        has_mdm = any("jumpcloud" in d.get("sources", []) for d in non_acked)
        statuses = list({d.get("status") for d in non_acked})

        people.append({
            "email": email,
            "device_count": total_count,
            "managed_count": managed_count,
            "has_edr": has_edr,
            "has_mdm": has_mdm,
            "compliant": managed_count > 0,
            "statuses": statuses,
            "devices": [{
                "hostname": (d.get("hostnames") or ["?"])[0],
                "status": d.get("status"),
                "sources": d.get("sources"),
                "serial": d.get("serial_number"),
                "os": d.get("os_type"),
                "confidence": d.get("confidence_score"),
            } for d in devs],
        })

    compliant = sum(1 for p in people if p["compliant"])
    return JSONResponse(content={
        "total_people": len(people),
        "compliant": compliant,
        "non_compliant": len(people) - compliant,
        "people": people,
    })


@app.get("/api/user/{email}/compliance")
async def api_user_compliance(email: str) -> Any:
    """Check if a user has at least one managed device."""
    repo = _get_repo()
    devices = repo.get_all_devices()
    acked = repo.get_acknowledged()

    user_devices = [d for d in devices
                    if (d.get("owner_email") or "").lower() == email.lower()
                    and d.get("canonical_id") not in acked]

    if not user_devices:
        return JSONResponse(content={"email": email, "found": False, "compliant": False, "devices": []})

    managed = [d for d in user_devices if d.get("status") in ("MANAGED", "FULLY_MANAGED")]
    has_edr = any("crowdstrike" in d.get("sources", []) for d in user_devices)
    has_mdm = any("jumpcloud" in d.get("sources", []) for d in user_devices)

    return JSONResponse(content={
        "email": email,
        "found": True,
        "compliant": len(managed) > 0,
        "has_edr": has_edr,
        "has_mdm": has_mdm,
        "device_count": len(user_devices),
        "managed_count": len(managed),
        "devices": [{
            "hostname": (d.get("hostnames") or ["?"])[0],
            "status": d.get("status"),
            "sources": d.get("sources"),
        } for d in user_devices],
    })


@app.get("/api/dual-use")
async def api_dual_use() -> Any:
    """People using both corporate and personal devices."""
    repo = _get_repo()
    devices = repo.get_all_devices()

    # Group by owner email
    by_owner: dict[str, list[dict[str, Any]]] = {}
    for d in devices:
        email = d.get("owner_email")
        if email:
            by_owner.setdefault(email.lower(), []).append(d)

    dual_users = []
    for email, devs in by_owner.items():
        corporate = [d for d in devs if "jumpcloud" in d.get("sources", []) or "crowdstrike" in d.get("sources", [])]
        personal = [d for d in devs if d.get("status") == "IDP_ONLY"]
        if corporate and personal:
            dual_users.append({
                "email": email,
                "corporate_devices": [{
                    "hostname": (d.get("hostnames") or ["?"])[0],
                    "status": d.get("status"),
                    "sources": d.get("sources"),
                    "serial": d.get("serial_number"),
                } for d in corporate],
                "personal_devices": [{
                    "hostname": (d.get("hostnames") or ["?"])[0],
                    "status": d.get("status"),
                    "serial": d.get("serial_number"),
                    "os": d.get("os_type"),
                } for d in personal],
            })

    return JSONResponse(content={
        "dual_use_count": len(dual_users),
        "total_users_with_devices": len(by_owner),
        "users": dual_users,
    })


@app.get("/api/insights")
async def api_insights() -> Any:
    if "insights" in _cache:
        return JSONResponse(content=_cache["insights"])
    repo = _get_repo()
    devices = repo.get_all_devices()
    summary = repo.get_summary()
    history = repo.get_status_history(limit=10)
    actions = generate_insights(devices, summary, history)
    return JSONResponse(content={"actions": actions})


@app.get("/api/report")
async def api_report() -> Any:
    repo = _get_repo()
    devices = repo.get_all_devices()
    summary = repo.get_summary()
    history = repo.get_status_history(limit=10)
    text = generate_report_text(devices, summary, history)
    return JSONResponse(content={"report": text})


@app.get("/api/report/full")
async def api_report_full() -> Any:
    """Full structured report data for PDF generation."""
    repo = _get_repo()
    devices = repo.get_all_devices()
    summary = repo.get_summary()
    # Add risk score
    by_status = summary.get("by_status", {})
    total_devices = summary.get("total", 0)
    if total_devices > 0:
        weighted = sum(by_status.get(s, 0) * w for s, w in RISK_WEIGHTS.items())
        summary["risk_score"] = round(weighted / total_devices, 1)
    else:
        summary["risk_score"] = 0
    history = repo.get_status_history(limit=10)
    last_sync = repo.get_last_sync_run()

    # AI executive summary
    report_text = generate_report_text(devices, summary, history)

    # Quick actions
    actions = generate_insights(devices, summary, history)

    # Helper to pick fields for PDF lists
    def device_summary(d: dict[str, Any]) -> dict[str, Any]:
        return {
            "hostname": (d.get("hostnames") or ["N/A"])[0],
            "serial": d.get("serial_number") or "N/A",
            "owner": d.get("owner_email") or "N/A",
            "os": d.get("os_type") or "N/A",
            "sources": d.get("sources", []),
            "status": d.get("status", "UNKNOWN"),
            "confidence": d.get("confidence_score", 0),
            "match_reason": d.get("match_reason", ""),
            "days_since_seen": d.get("days_since_seen"),
        }

    # Top devices per category
    no_edr = [device_summary(d) for d in devices if d.get("status") == "NO_EDR"][:15]
    no_mdm = [device_summary(d) for d in devices if d.get("status") == "NO_MDM"][:15]
    idp_only = [device_summary(d) for d in devices if d.get("status") == "IDP_ONLY"][:15]
    stale = sorted(
        [device_summary(d) for d in devices if d.get("status") == "STALE"],
        key=lambda x: x.get("days_since_seen") or 0, reverse=True,
    )[:10]

    # Unique devices with match explanation (high confidence multi-source)
    unique_matches = []
    for d in devices:
        sources = d.get("sources", [])
        if len(sources) >= 2 and (d.get("confidence_score") or 0) >= 0.6:
            unique_matches.append(device_summary(d))
    unique_matches.sort(key=lambda x: x["confidence"], reverse=True)

    # Low confidence (potential duplicates or bad matches)
    low_confidence = sorted(
        [device_summary(d) for d in devices if (d.get("confidence_score") or 0) < 0.5],
        key=lambda x: x["confidence"],
    )[:15]

    return JSONResponse(content={
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "executive_summary": report_text,
        "summary": summary,
        "last_sync": last_sync,
        "actions": actions,
        "categories": {
            "no_edr": {"title": "Devices without EDR (CrowdStrike)", "count": len([d for d in devices if d.get("status") == "NO_EDR"]), "devices": no_edr},
            "no_mdm": {"title": "Devices without MDM (JumpCloud)", "count": len([d for d in devices if d.get("status") == "NO_MDM"]), "devices": no_mdm},
            "idp_only": {"title": "IDP-Only Devices (potential shadow IT)", "count": len([d for d in devices if d.get("status") == "IDP_ONLY"]), "devices": idp_only},
            "stale": {"title": "Stale Devices (90+ days inactive)", "count": len([d for d in devices if d.get("status") == "STALE"]), "devices": stale},
        },
        "unique_matches": {"title": "Cross-source Matched Devices", "count": len(unique_matches), "devices": unique_matches[:20]},
        "low_confidence": {"title": "Low Confidence Matches (review needed)", "count": len(low_confidence), "devices": low_confidence},
    })


@app.post("/api/slack/test")
async def api_slack_test() -> Any:
    """Send a test Slack alert with current data using Block Kit."""
    from src.alerts import send_slack, build_sync_blocks, _get_webhook_url
    if not _get_webhook_url():
        return JSONResponse(content={"error": "SLACK_WEBHOOK_URL not configured in .env"}, status_code=400)

    repo = _get_repo()
    devices = repo.get_all_devices()
    summary_data = repo.get_summary()
    by_status = summary_data.get("by_status", {})
    total = summary_data.get("total", 0)
    managed = (by_status.get("MANAGED", 0) + by_status.get("FULLY_MANAGED", 0))
    no_edr = sum(1 for d in devices if d.get("status") == "NO_EDR")
    no_mdm = sum(1 for d in devices if d.get("status") == "NO_MDM")

    disappeared = repo.get_recently_deleted()
    newly_stale = repo.get_newly_stale()

    blocks = build_sync_blocks(
        status_counts=by_status,
        total=total,
        managed=managed,
        sources_ok=["crowdstrike", "jumpcloud", "okta"],
        sources_failed=[],
        sync_status="test",
        disappeared=disappeared,
        newly_stale=newly_stale,
        no_edr_count=no_edr,
        no_mdm_count=no_mdm,
    )

    fallback = f"Klar Test: {total} devices, {managed} managed"
    ok = send_slack(fallback, blocks=blocks)
    return JSONResponse(content={"sent": ok, "blocks_count": len(blocks)})


class TriggerResponse(BaseModel):
    message: str
    started: bool


@app.post("/api/sync/trigger")
async def api_sync_trigger(background_tasks: BackgroundTasks) -> Any:
    def _run() -> None:
        try:
            SyncEngine(DB_PATH).run()
            refresh_cache()
        except Exception:
            pass

    background_tasks.add_task(_run)
    return JSONResponse(content={"message": "Sync triggered", "started": True})


# ── SPA Catch-all ────────────────────────────────────────────────────────────

@app.get("/")
async def serve_index(request: Request) -> Any:
    index_path = DIST_DIR / "index.html"
    if index_path.exists():
        return FileResponse(index_path)
    return JSONResponse(
        {"detail": "Frontend not built. Run: cd frontend && npm run build"},
        status_code=404,
    )


@app.get("/{path:path}")
async def serve_spa(path: str, request: Request) -> Any:
    # API 404s should stay JSON
    if path.startswith("api/"):
        return JSONResponse({"detail": "Not found"}, status_code=404)

    # Try to serve static file directly
    # Resolve to absolute path and verify it stays within DIST_DIR to prevent
    # path traversal (e.g. GET /../../etc/passwd).  resolve() follows .. and
    # symlinks; we check the canonical prefix before serving the file.
    file_path = (DIST_DIR / path).resolve()
    try:
        file_path.relative_to(DIST_DIR.resolve())
    except ValueError:
        return JSONResponse({"detail": "Not found"}, status_code=404)
    if file_path.exists() and file_path.is_file():
        return FileResponse(file_path)

    # Fallback to SPA index.html
    index_path = DIST_DIR / "index.html"
    if index_path.exists():
        return FileResponse(index_path)

    return JSONResponse(
        {"detail": "Frontend not built. Run: cd frontend && npm run build"},
        status_code=404,
    )
